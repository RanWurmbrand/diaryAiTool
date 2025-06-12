import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font,Border, Side,Alignment
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import range_boundaries
import os

header_txt = "Intern Instructions: The work diary is meant to help you track your progress over the summer, reflect on your accomplishments, and support conversations with your manager, team buddy, or mentor. It will also serve as a\n helpful reference when you complete a competency document later in your internship. Please remove examples below and fill in your own entries weekly, this should take no more than 15 minutes each week."
meta_data_path = "metadataxlsx"

files_path = "xlsxfiles"


background_color_pallet = {
    "yellow": "FFF700",
    "red":    "FA7B5C",
    "grey":   "A199A8",
    "blue":   "5FBCFA",
    "green":  "8EE46F" 
}
text_color_pallet = {
    "yellow": "FFFF00",
    "red":    "FF0000",
    "blue":   "0000FF",
    "green":  "00FF00" 
}

bold_border =Border(
    left = Side(style='thick'),
    right =Side(style='thick'),
    top= Side(style='thick'),
    bottom= Side(style='thick')
)
def is_xlsx_exist(path):
    if not path.endswith(".xlsx"):
        return False
    return os.path.exists(path)

class XlsxTool:


    def __init__(self, file_name):  # Fixed: __init__ instead of **init**
        if len(file_name) == 0 or (not file_name.endswith(".xlsx")):
            print("invalid filename")
            return None
        self.fName = f"{files_path}/{file_name}"
        file_path = f"{files_path}/{file_name}"
        if os.path.exists(file_path):
            # Load existing file to preserve content
            self.wb = openpyxl.load_workbook(file_path)
        else:
            # Create new workbook only if file doesn't exist
            self.wb = Workbook()
    

    def save_file(self):
        self.wb.save(self.fName)
    
    def create_file(self,file_name,max_width = None):
        file_path = f"{files_path}/{file_name}"
        if os.path.exists(file_path):
            print("file already exist!")
            return None
        else:
            os.makedirs(files_path, exist_ok=True)  # Create directory if it doesn't exist


    def change_row_size(self,row_num,size = 0):
        if size <= 0:
            print("invalid size")
            return
        ws = self.wb.active
        ws.row_dimensions[row_num].height = size


    def add_text(self,text,row,col):
        ws = self.wb.active
        ws.cell(row=row, column= col,value=text)



    def add_write_bold_text(self,text,row,col):
        bold_font = Font(bold=True)
        ws = self.wb.active
        cell = ws.cell(row=row, column= col,value=text)
        cell.font = bold_font


    def add_colored_text(self,text,row,col,color):
        rgb_color = "000000"
        color = color.lower()
        if color in text_color_pallet:
            rgb_color = text_color_pallet[color]
        color_font = Font(color = rgb_color)
        ws = self.wb.active
        cell = ws.cell(row=row, column= col,value=text)
        cell.font = color_font



    def add_two_colored_text_sections(self,row,col,first_color,first_text,second_color,second_text):
        first_rgb_color = "000000"
        second_rgb_color = "000000"
        first_color = first_color.lower()
        second_color = second_color.lower()
        first_text += " "
        if first_color in text_color_pallet:
            first_rgb_color = text_color_pallet[first_color]
        if second_color in text_color_pallet:
            second_rgb_color = text_color_pallet[second_color]
        print(first_rgb_color,first_text)
        first_section = TextBlock(InlineFont(color=first_rgb_color),first_text)
        second_section = TextBlock(InlineFont(color=second_rgb_color),second_text)
        rich_text = CellRichText(first_section,second_section)
        ws = self.wb.active
        cell = ws.cell(row=row, column=col, value=rich_text) 




    def add_three_colored_text_sections(self,row,col,first_color,first_text,second_color,second_text,third_color,third_text):
        first_rgb_color = "000000"
        second_rgb_color = "000000"
        third_rgb_color = "000000"
        first_color = first_color.lower()
        second_color = second_color.lower()
        third_color = third_color.lower()
        first_text += " "
        second_text += " "
        if first_color in text_color_pallet:
            first_rgb_color = text_color_pallet[first_color]
        if second_color in text_color_pallet:
            second_rgb_color = text_color_pallet[second_color]
        if third_color in text_color_pallet:
            third_rgb_color = text_color_pallet[third_color]
        
        first_section = TextBlock(InlineFont(color=first_rgb_color),first_text)
        second_section = TextBlock(InlineFont(color=second_rgb_color),second_text)
        third_section = TextBlock(InlineFont(color=third_rgb_color),third_text)
        rich_text = CellRichText(first_section,second_section,third_section)
        ws = self.wb.active
        cell = ws.cell(row=row, column=col, value=rich_text) 



    def add_hyper_link_text(self,row,col,link,placeholder = ""):
        if placeholder =="":
            placeholder = link
        ws = self.wb.active
        cell = ws.cell(row=row,column=col,value=placeholder)
        cell.hyperlink = link
        cell.font =Font(color=text_color_pallet["blue"],underline="single")



    def add_options_cell(self,row,col,options):
        if 0 == len(options):
            print("no options given")
            return
        str_options = ','.join(options)
        dv = DataValidation(type="list",formula1=f'"{str_options}"') 
        dv.showErrorMessage = False
        
        ws = self.wb.active
        cell = ws.cell(row=row,column = col)
        dv.add(cell)
        cell.value=options[0]


    
    # def add_sorting_cell(self,row,col):
    #     pass


    def make_cell_bold(self,row,col):
        if row < 0 or col < 0:
            print("invalid input")
        
        ws = self.wb.active
        ws = self.wb.active
        cell = ws.cell(row= row, column = col)
        cell.border = bold_border

    def merged_cells(self,start_row,start_col,end_row,end_col):
        if start_row < 0 or start_col < 0 or end_row < start_row or end_col < start_col:
            print("invalid input!")
            return
        ws = self.wb.active
        ws.merge_cells(start_row = start_row, start_column = start_col, end_row = end_row,end_column = end_col)
        
    def make_mereged_cells_bold(self,start_row,start_col):
        ws = self.wb.active
        cell_coord = ws.cell(row=start_row,column= start_col).coordinate
        cell_coord_str = str(cell_coord)
        merged_range_str = ""
        for merged_range in ws.merged_cells.ranges:
            print(merged_range,cell_coord)
            if cell_coord in merged_range:
                merged_range_str =str(merged_range)
                break
        if merged_range_str == "":
            print("cell are not mereged")
            return
        min_col, min_row, max_col, max_row = range_boundaries(merged_range_str)
        
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row=row, column=col).border = bold_border
                
    def add_divided_merged_cells(self,start_row,start_col,mereged_size,num_of_cells):
        for i in range(num_of_cells):
            self.merged_cells(start_row,start_col,(start_row + i*mereged_size),start_col + mereged_size)


    def dye_row(self,row_num,row_width):
        pass

    def add_merged_row_with_hight(self,start,end):
        pass


    def add_options_row(self,options):
        pass
    

    # TODO safty check that we dont override existing data 
    def add_matrix(self,start_row,start_col,width,height):
        for i in range(height + 1):
            for j in range(width + 1):
                self.make_cell_bold(start_row + i, start_col + j)

    def add_titled_matrix(self,start_row,start_col,width,height,row_titles,col_titles):
        pass

    def add_side_titled_matrix(self,start_row,start_col,width,height,titles):
        self.add_matrix(start_row,start_col,width,height)
        ws = self.wb.active
        n = min(height + 1, len(titles))
        for i in range(n):
            cell = ws.cell(row= start_row + i,column = start_col,value =titles[i])
            cell.alignment = Alignment(horizontal='center', vertical='center')

    def add_top_titled_matrix(self,start_row,start_col,width,height,titles):
        self.add_matrix(start_row,start_col,width,height)
        ws = self.wb.active
        n = min(width + 1, len(titles))
        for i in range(n):
            cell = ws.cell(row= start_row,column = start_col + i,value =titles[i])
            cell.alignment = Alignment(horizontal='center', vertical='center')